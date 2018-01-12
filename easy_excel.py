import openpyxl

# Open Workbook
def open_workbook(file_path):
    try:
        return openpyxl.load_workbook(file_path)
    except:
        print("Open Workbook Error: " + file_path)
        return None
    
# Open Sheet
def open_sheet(workbook, index):
    try:
        sheet_name = workbook.get_sheet_names()[index]
        return workbook.get_sheet_by_name(sheet_name)
    except:
        print("Open Sheet Error: " + str(workbook) + " " + str(index))

# Create Sheet
def create_sheet(workbook, sheet_name):
    try:
        sheets = workbook.get_sheet_names()
        if sheet_name in sheets:
            return workbook.get_sheet_by_name(sheet_name)
        else:
            return workbook.create_sheet(title = sheet_name)
    except:
        print("Create Sheet Error: " + sheet_name)
        
# Get Sheet Data
def get_data(sheet):
    data = []
    try:
        row_max = sheet.max_row
        col_max = sheet.max_column
        for row in sheet.iter_rows(min_row = 2, max_col = col_max, max_row = row_max):
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            data.append(row_values)
    except:
        print("Data Retireve Error " + str(sheet))
    return data

# Fill Sheet with Data
def fill_sheet(sheet, titles, data):
    x = 1
    for y in range(0, len(titles)):
        x += 1
        sheet.cell(row = 1, column = y + 1).value = titles[y]
        print(titles[y])
        for row in data:
            sheet.cell(row = x, column = y + 1).value = row[y]
        
