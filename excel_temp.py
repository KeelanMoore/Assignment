# Normal Imports
import openpyxl
import sqlite3

# Custom Imports
from easy_sql import *

# Create Excel File
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Temperature by city"

# Query
cursor.execute("""SELECT city, av_temp
                  FROM city_table
                  WHERE country LIKE 'China' and av_temp is NOT NULL
                  ORDER BY city""")

data = cursor.fetchall()
new_data = {}

for values in data:
    city = values[0]
    temp = values[1]
    if city in new_data:
        new_data[city][0] += temp
        new_data[city][1] += 1
    else:
        new_data[city] = [temp,1]

city_means = []
for city in new_data:
    mean = new_data[city][0] / new_data[city][1]
    city_means.append([city, mean])

i = 0
for values in city_means:
    i += 1
    sheet.cell(row = i, column = 1).value = values[0]
    sheet.cell(row = i, column = 2).value = values[1]

# Save
file_path = "World Temperature.xlsx"
workbook.save(filename = file_path)
